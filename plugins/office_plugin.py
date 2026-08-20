# office_mod.py - 通用办公插件
import os
import re
from datetime import datetime
from typing import Dict, List
from plugin_base import PluginBase

_openpyxl_loaded = False


def _ensure_openpyxl():
    """惰性导入：首次生成 Excel 时才加载 openpyxl"""
    global _openpyxl_loaded, Workbook, Font, Alignment, PatternFill, get_column_letter
    if not _openpyxl_loaded:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        _openpyxl_loaded = True


def _openpyxl_available():
    try:
        import importlib.util
        return importlib.util.find_spec("openpyxl") is not None
    except Exception:
        return False


class OfficeMod(PluginBase):
    name = "通用办公"
    version = "1.0"
    description = "学校、工作、会计场景Excel模板生成"
    author = "seiki"
    enabled = True

    def __init__(self, core):
        super().__init__(core)
        self.export_dir = "exports"
        os.makedirs(self.export_dir, exist_ok=True)

        self.scenes = {
            "school": {
                "name": "学校",
                "templates": {
                    "class": self._school_class,
                    "grade": self._school_grade,
                    "student": self._school_student,
                    "exam": self._school_exam,
                    "leave": self._school_leave,
                }
            },
            "work": {
                "name": "工作",
                "templates": {
                    "meeting": self._work_meeting,
                    "daily": self._work_daily,
                    "weekly": self._work_weekly,
                    "staff": self._work_staff,
                    "approval": self._work_approval,
                    "reimburse": self._work_reimburse,
                }
            },
            "accounting": {
                "name": "会计",
                "templates": {
                    "income": self._acc_income,
                    "invoice": self._acc_invoice,
                    "budget": self._acc_budget,
                    "reconcile": self._acc_reconcile,
                    "asset": self._acc_asset,
                }
            },
            "general": {
                "name": "通用",
                "templates": {
                    "blank": self._gen_blank,
                    "table": self._gen_table,
                }
            }
        }

    def on_command(self, command, args):
        if not _openpyxl_available():
            return "请安装 openpyxl: pip install openpyxl", False

        if command == "office":
            return self._handle(args), False
        elif command == "office_help":
            return self._help(), False
        return None

    def _help(self):
        lines = [
            "/office <场景> <模板> [参数]",
            "",
            "【school】 class(课程表) grade(成绩表) student(学生名单) exam(考试安排) leave(请假条)",
            "【work】 meeting(会议纪要) daily(日报) weekly(周报) staff(人员统计) approval(审批表) reimburse(报销单)",
            "【accounting】 income(收支明细) invoice(发票登记) budget(预算表) reconcile(对账单) asset(固定资产)",
            "【general】 blank(空白表格) table(数据表)",
            "",
            "示例: /office school class 班级=2024级1班",
        ]
        return "\n".join(lines)

    def _handle(self, args):
        _ensure_openpyxl()
        parts = args.strip().split()
        if len(parts) < 2:
            return self._help()

        scene_key, template_key = parts[0].lower(), parts[1].lower()
        data = self._parse(" ".join(parts[2:]))

        scene = self.scenes.get(scene_key)
        if not scene:
            return f"未知场景: {scene_key}\n{self._help()}"

        func = scene["templates"].get(template_key)
        if not func:
            return f"未知模板: {template_key}\n{self._help()}"

        try:
            return func(data)
        except Exception as e:
            return f"生成失败: {e}"

    def _parse(self, raw):
        result = {}
        for m in re.finditer(r'(\w+)="(.*?)"', raw):
            result[m.group(1)] = m.group(2)
        for m in re.finditer(r'(\w+)=(\S+)', raw):
            if m.group(1) not in result:
                result[m.group(1)] = m.group(2)
        return result

    def _save(self, wb, name):
        try:
            import doc_layout
            doc_layout.apply_excel_layout(wb)  # 自动列宽 + 表头样式 + 边框
        except Exception as e:
            print(f"[通用办公] 排版失败: {e}")
        fn = f"{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = os.path.join(self.export_dir, fn)
        wb.save(path)
        return f"✅ {name}已生成: {path}"

    def _title(self, ws, text):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        c = ws.cell(row=1, column=1, value=text)
        c.font = Font(name="宋体", size=16, bold=True)
        c.alignment = Alignment(horizontal="center")

    def _headers(self, ws, headers, row=3):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
            c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # ========== school ==========
    def _school_class(self, d):
        wb, ws = Workbook(), Workbook().active
        ws.title = "课程表"
        self._title(ws, f"{d.get('班级','')} 课程表")
        days = ["星期", "周一", "周二", "周三", "周四", "周五"]
        for i, v in enumerate(days, 1):
            ws.cell(row=3, column=i, value=v).font = Font(bold=True)
        periods = d.get("时段", "第1节,第2节,第3节,第4节,第5节,第6节,第7节").split(',')
        for i, p in enumerate(periods, 4):
            ws.cell(row=i, column=1, value=p).font = Font(bold=True)
        ws.column_dimensions['A'].width = 12
        for i in range(2, 7):
            ws.column_dimensions[get_column_letter(i)].width = 14
        return self._save(wb, f"课程表_{d.get('班级','')}")

    def _school_grade(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "成绩表"
        self._title(ws, f"{d.get('班级','')} 成绩表")
        subs = ["序号", "姓名", "语文", "数学", "英语"] + [s.strip() for s in d.get("科目", "").split(',') if s.strip()] + ["总分", "排名"]
        self._headers(ws, subs)
        for i in range(4, 4 + int(d.get("人数", "5"))):
            ws.cell(row=i, column=1, value=i-3)
            ws.cell(row=i, column=2, value=f"学生{i-3}")
        for i in range(1, len(subs)+1):
            ws.column_dimensions[get_column_letter(i)].width = 12
        return self._save(wb, f"成绩表_{d.get('班级','')}")

    def _school_student(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "学生名单"
        self._title(ws, f"{d.get('班级','')} 学生名单")
        self._headers(ws, ["序号", "姓名", "性别", "学号", "家长电话", "备注"])
        for i in range(4, 4 + int(d.get("人数", "10"))):
            ws.cell(row=i, column=1, value=i-3)
        for i, w in enumerate([8, 12, 10, 16, 18, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return self._save(wb, f"学生名单_{d.get('班级','')}")

    def _school_exam(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "考试安排"
        self._title(ws, d.get("标题", "考试安排"))
        self._headers(ws, ["日期", "时间", "科目", "考场", "监考老师", "备注"])
        exams = d.get("科目列表", "语文,数学,英语,物理,化学").split(',')
        for i, e in enumerate(exams, 1):
            row = 3 + i
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=3, value=e.strip())
        for i, w in enumerate([12, 18, 15, 15, 18, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return self._save(wb, "考试安排")

    def _school_leave(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "请假条"
        self._title(ws, "请假条")
        for label, key in [("请假人", "姓名"), ("班级", "班级"), ("请假时间", "时间"), ("请假天数", "天数")]:
            ws.cell(row=3 + len(ws['A']), column=1, value=label).font = Font(bold=True)
            ws.cell(row=3 + len(ws['A']), column=2, value=d.get(key, ""))
        r = ws.max_row + 2
        ws.cell(row=r, column=1, value="班主任签字：")
        ws.cell(row=r, column=2, value="日期：")
        ws.cell(row=r+2, column=1, value="家长签字：")
        ws.cell(row=r+2, column=2, value="日期：")
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 40
        return self._save(wb, f"请假条_{d.get('姓名','')}")

    # ========== work ==========
    def _work_meeting(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "会议纪要"
        self._title(ws, d.get("标题", "会议纪要"))
        for label, key in [("会议时间", "时间"), ("会议地点", "地点"), ("主持人", "主持"), ("记录人", "记录"), ("参会人员", "参会")]:
            ws.cell(row=3 + len(ws['A']), column=1, value=label).font = Font(bold=True)
            ws.cell(row=3 + len(ws['A']), column=2, value=d.get(key, ""))
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value="会议议题").font = Font(bold=True)
        ws.cell(row=r+1, column=1, value=d.get("议题", "")).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=2)
        r += 3
        ws.cell(row=r, column=1, value="会议决议").font = Font(bold=True)
        ws.cell(row=r+1, column=1, value=d.get("决议", "")).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=2)
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 40
        return self._save(wb, "会议纪要")

    def _work_daily(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "工作日报"
        self._title(ws, f"{d.get('姓名','')} 日报 {datetime.now().strftime('%Y-%m-%d')}")
        self._headers(ws, ["序号", "工作内容", "完成情况", "问题建议", "明日计划"])
        for i in range(4, 7):
            ws.cell(row=i, column=1, value=i-3)
        for i, w in enumerate([8, 30, 15, 25, 25], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        r = ws.max_row + 2
        ws.cell(row=r, column=1, value=f"填报人：{d.get('姓名','')}")
        ws.cell(row=r+1, column=1, value=f"日期：{datetime.now().strftime('%Y-%m-%d')}")
        return self._save(wb, f"日报_{d.get('姓名','')}")

    def _work_weekly(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "工作周报"
        self._title(ws, f"{d.get('姓名','')} 周报 ({d.get('周期','')})")
        self._headers(ws, ["序号", "本周工作", "完成情况", "下周计划", "备注"])
        for i in range(4, 9):
            ws.cell(row=i, column=1, value=i-3)
        for i, w in enumerate([8, 35, 20, 30, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        r = ws.max_row + 2
        ws.cell(row=r, column=1, value=f"填报人：{d.get('姓名','')}")
        ws.cell(row=r+1, column=1, value=f"日期：{datetime.now().strftime('%Y-%m-%d')}")
        return self._save(wb, f"周报_{d.get('姓名','')}")

    def _work_staff(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "人员统计"
        self._title(ws, d.get("标题", "人员信息统计表"))
        self._headers(ws, ["序号", "姓名", "部门", "职位", "入职日期", "联系方式"])
        cnt = int(d.get("人数", "10"))
        for i in range(4, 4 + cnt):
            ws.cell(row=i, column=1, value=i-3)
        for i, w in enumerate([8, 12, 15, 15, 18, 18], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        r = 4 + cnt + 2
        ws.cell(row=r, column=1, value="合计：").font = Font(bold=True)
        ws.cell(row=r, column=3, value=f"{cnt} 人")
        return self._save(wb, "人员统计")

    def _work_approval(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "审批单"
        self._title(ws, "审批申请表")
        for label, key in [("申请人", "申请人"), ("申请部门", "部门"), ("申请日期", "日期"), ("审批类别", "类别")]:
            ws.cell(row=3 + len(ws['A']), column=1, value=label).font = Font(bold=True)
            ws.cell(row=3 + len(ws['A']), column=2, value=d.get(key, ""))
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value="申请事由：").font = Font(bold=True)
        ws.cell(row=r+1, column=1, value=d.get("事由", "")).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=2)
        r += 3
        ws.cell(row=r, column=1, value="审批意见：").font = Font(bold=True)
        ws.cell(row=r+2, column=1, value="审批人签字：")
        ws.cell(row=r+2, column=2, value="日期：")
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 40
        return self._save(wb, f"审批单_{d.get('申请人','')}")

    def _work_reimburse(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "报销单"
        self._title(ws, "费用报销单")
        for label, key in [("报销人", "申请人"), ("部门", "部门"), ("报销日期", "日期")]:
            ws.cell(row=3 + len(ws['A']), column=1, value=label).font = Font(bold=True)
            ws.cell(row=3 + len(ws['A']), column=2, value=d.get(key, ""))
        r = ws.max_row + 2
        ws.cell(row=r, column=1, value="报销事由：").font = Font(bold=True)
        ws.cell(row=r+1, column=1, value=d.get("事由", ""))
        ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=2)
        r += 3
        self._headers(ws, ["序号", "费用项目", "费用明细", "金额", "备注"], row=r)
        for i in range(1, 6):
            ws.cell(row=r+i, column=1, value=i)
        for i, w in enumerate([8, 20, 25, 15, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return self._save(wb, f"报销单_{d.get('申请人','')}")

    # ========== accounting ==========
    def _acc_income(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "收支明细"
        self._title(ws, f"收支明细 ({d.get('月份', datetime.now().strftime('%Y年%m月'))})")
        self._headers(ws, ["日期", "摘要", "收入", "支出", "余额", "备注"])
        for i in range(4, 14):
            ws.cell(row=i, column=1, value=f"2026-08-{i-3:02d}")
        for i, w in enumerate([15, 25, 15, 15, 15, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        r = 15
        ws.cell(row=r, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=r, column=3, value="=SUM(C4:C13)")
        ws.cell(row=r, column=4, value="=SUM(D4:D13)")
        ws.cell(row=r, column=5, value="=C14-D14")
        return self._save(wb, f"收支明细_{d.get('月份','')}")

    def _acc_invoice(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "发票登记"
        self._title(ws, "发票登记表")
        self._headers(ws, ["序号", "发票号码", "开票日期", "金额", "税率", "税额", "对方单位", "备注"])
        for i in range(4, 14):
            ws.cell(row=i, column=1, value=i-3)
        for i, w in enumerate([8, 18, 15, 15, 12, 15, 25, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        r = 15
        ws.cell(row=r, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=r, column=4, value="=SUM(D4:D13)")
        ws.cell(row=r, column=6, value="=SUM(F4:F13)")
        return self._save(wb, "发票登记表")

    def _acc_budget(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "预算表"
        self._title(ws, f"预算表 ({d.get('年度', datetime.now().strftime('%Y年'))})")
        self._headers(ws, ["项目", "预算金额", "已支出", "剩余", "执行率", "备注"])
        items = d.get("项目", "办公费,差旅费,会议费,培训费,设备费").split(',')
        for i, item in enumerate(items, 4):
            ws.cell(row=i, column=1, value=item.strip())
        for i, w in enumerate([20, 15, 15, 15, 15, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        r = 4 + len(items) + 2
        ws.cell(row=r, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=r, column=2, value="=SUM(B4:B13)")
        ws.cell(row=r, column=3, value="=SUM(C4:C13)")
        return self._save(wb, f"预算表_{d.get('年度','')}")

    def _acc_reconcile(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "对账单"
        self._title(ws, f"对账单 ({d.get('月份', datetime.now().strftime('%Y年%m月'))})")
        for label, key in [("单位", "单位"), ("对账期间", "月份"), ("联系人", "联系人")]:
            ws.cell(row=3 + len(ws['A']), column=1, value=label).font = Font(bold=True)
            ws.cell(row=3 + len(ws['A']), column=2, value=d.get(key, ""))
        r = ws.max_row + 1
        self._headers(ws, ["序号", "业务日期", "业务内容", "对方金额", "我方金额", "差异"], row=r)
        for i in range(1, 11):
            ws.cell(row=r+i, column=1, value=i)
        for i, w in enumerate([8, 15, 25, 15, 15, 15], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return self._save(wb, f"对账单_{d.get('单位','')}")

    def _acc_asset(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "固定资产"
        self._title(ws, d.get("标题", "固定资产登记表"))
        self._headers(ws, ["序号", "资产名称", "规格型号", "数量", "单价", "总价", "存放地点", "使用人", "备注"])
        for i in range(4, 14):
            ws.cell(row=i, column=1, value=i-3)
        for i, w in enumerate([8, 15, 18, 10, 12, 15, 15, 12, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        r = 15
        ws.cell(row=r, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=r, column=6, value="=SUM(F4:F13)")
        return self._save(wb, "固定资产表")

    # ========== general ==========
    def _gen_blank(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "空白表格"
        self._title(ws, d.get("标题", "空白表格"))
        rows, cols = int(d.get("行数", "10")), int(d.get("列数", "5"))
        for i in range(1, cols+1):
            ws.cell(row=3, column=i, value=f"列{i}").font = Font(bold=True)
        for i in range(1, cols+1):
            ws.column_dimensions[get_column_letter(i)].width = 15
        return self._save(wb, "空白表格")

    def _gen_table(self, d):
        wb = Workbook()
        ws = wb.active
        ws.title = "数据表"
        self._title(ws, d.get("标题", "数据汇总表"))
        headers = [h.strip() for h in d.get("表头", "项目,数值1,数值2,数值3").split(',')]
        self._headers(ws, headers)
        for i in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(i)].width = 18
        return self._save(wb, "数据表")