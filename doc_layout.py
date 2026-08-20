# ============================================================
#   doc_layout.py - 排版算法
#
#   Word：按党政机关公文格式 GB/T 9704 自动排版
#     · 页边距：上 3.7cm 下 3.5cm 左 2.8cm 右 2.6cm
#     · 标题：2号（22pt）小标宋，居中
#     · 一级标题（一、二、…）：3号黑体加粗
#     · 正文：3号（16pt）仿宋_GB2312，首行缩进 2 字符，行距固定 28 磅
#     · 落款/日期：右对齐
#   Excel：自动列宽（中文按 2 字符宽计）、表头加粗居中、全表细边框
# ============================================================

from docx.enum.text import WD_ALIGN_PARAGRAPH, WD_LINE_SPACING
from docx.oxml.ns import qn
from docx.shared import Cm, Pt


def _set_run_font(run, zh_font, size_pt, bold=False):
    """设置中西文字体（西文 Times New Roman + 中文字体）"""
    run.font.name = "Times New Roman"
    run.font.size = Pt(size_pt)
    run.font.bold = bold
    rpr = run._element.get_or_add_rPr()
    rfonts = rpr.get_or_add_rFonts()
    rfonts.set(qn("w:eastAsia"), zh_font)


def _is_head1(text):
    return len(text) >= 3 and text[0] in "一二三四五六七八九十" and text[1] == "、"


def apply_official_format(doc, title_lines=1):
    """把 Word 文档按公文规范排版（启发式：前 title_lines 个非空段视为标题）"""
    sec = doc.sections[0]
    sec.top_margin = Cm(3.7)
    sec.bottom_margin = Cm(3.5)
    sec.left_margin = Cm(2.8)
    sec.right_margin = Cm(2.6)

    paras = doc.paragraphs
    total = len(paras)
    title_seen = 0
    for i, p in enumerate(paras):
        text = (p.text or "").strip()
        if not text:
            continue
        is_title = title_seen < title_lines
        is_sign = (i >= total - 2) and (
            "落款" in text or "单位" in text or "日期" in text
            or (len(text) < 25 and "：" not in text and not text.endswith("。")))
        if is_title:
            title_seen += 1
            p.alignment = WD_ALIGN_PARAGRAPH.CENTER
            for run in p.runs:
                _set_run_font(run, "方正小标宋简体", 22, bold=False)
        elif _is_head1(text):
            p.paragraph_format.first_line_indent = Pt(32)
            for run in p.runs:
                _set_run_font(run, "黑体", 16, bold=True)
        elif is_sign:
            p.alignment = WD_ALIGN_PARAGRAPH.RIGHT
            for run in p.runs:
                _set_run_font(run, "仿宋_GB2312", 16, bold=False)
        else:
            pf = p.paragraph_format
            pf.first_line_indent = Pt(32)
            pf.line_spacing_rule = WD_LINE_SPACING.EXACTLY
            pf.line_spacing = Pt(28)
            for run in p.runs:
                _set_run_font(run, "仿宋_GB2312", 16, bold=False)
    return doc


def apply_excel_layout(wb, max_width=60, min_width=10):
    """Excel 自动排版：列宽自适应（中文按 2 计）、表头加粗居中、细边框"""
    from openpyxl.styles import Alignment, Border, Font, Side

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    from openpyxl.utils import get_column_letter
    for ws in wb.worksheets:
        # 1) 自动列宽
        widths = {}
        for row in ws.iter_rows(values_only=True):
            for ci, v in enumerate(row, 1):
                if v is None:
                    continue
                w = sum(2 if ord(ch) > 127 else 1 for ch in str(v))
                widths[ci] = max(widths.get(ci, 0), w)
        for ci, w in widths.items():
            letter = get_column_letter(ci)
            ws.column_dimensions[letter].width = max(min_width, min(w + 2, max_width))
        # 2) 表头（第一行非空单元格）加粗居中
        for cell in ws[1]:
            if cell.value is not None:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        # 3) 数据区细边框
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = border
    return wb
